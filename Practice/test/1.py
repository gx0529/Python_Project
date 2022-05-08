
import openpyxl

# from openpyxl.cell import get_column_letter, column_index_from_string
from openpyxl.utils import get_column_letter, column_index_from_string

# wb = openpyxl.load_workbook("./example.xlsx")

# 取出excel文件中所有的表名
# sheet_name = wb.sheetnames

# 取出特定表名的sheet对象
# sheet = wb.get_sheet_by_name("Sheet1")



# print(sheet['A1'])
# print(sheet.title)
# print(sheet['A1'].value)
# Cell对象有一个value属性
# print(sheet.cell(row=1,column=2).value)

# for i in range(1,8,1):
#     print(sheet.cell(row=i,column=2).value)

# print(sheet.get_highest_column())
# get_highest_row()和get_highest_column()在最新版的openpyxl模块中已经被删除了，
# 取而代之的是max_row和max_column两个方法。
#
# 注意：使用的时候不用在方法后面添加括号
# print(sheet.max_row)

# 数字转换为字母函数
# print(get_column_letter(2))

# 字母转换为数字函数
# print(column_index_from_string("A"))

# 对sheet对象进行修改且保存
# sheet.title = "spam"
# wb.save("example.xlsx")

# 创建和删除工作表
wb = openpyxl.load_workbook(filename="./example.xlsx",data_only=True)
# wb.create_sheet()
# wb.create_sheet(index=1,title="second_sheet")
# wb.remove(wb.get_sheet_by_name("second_heet"))

sheet = wb.get_sheet_by_name("spam")
# sheet["A1"] = "hello world"
# print(sheet["A1"].value)

# sheet['C8'] = '=SUM(C1:C7)'
# sheet['C8'] = '=SUM(C1:C7)'
print(sheet['C8'].value)
wb.save("./example.xlsx")
test = wb.sheetnames
print(test)


