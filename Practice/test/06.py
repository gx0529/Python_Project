
import openpyxl,os

wb = openpyxl.Workbook()

sheet = wb.create_sheet("zonghe")

fileList = ["pc.txt"]

# ,"pc.txt","crc.txt"
# file = open("./apk.txt")
# fileLine = file.readline()
# print(len(fileLine))
# print(fileLine)
dir = os.getcwd()
for j in range(1,len(fileList)+1):
    file = open(os.path.join(dir,fileList[j-1]),encoding="UTF-8")
    fileLine = file.read().split("\n")
    pos = fileList[j-1].find(".")
    sheet.cell(row=1,column=j).value = fileList[j-1][:pos]
    for i in range(1,len(fileLine)):
        sheet.cell(row=i+1, column=j).value = str(fileLine[i-1])
        print(sheet.cell(row=i+1, column=j).value)
wb.save('./result.xlsx')
